# -*- coding: utf-8 -*-
"""
村户慧眼台账系统 — 数据生成器
解析脱敏 Excel -> 结构化「户-成员-产业/养殖-位置」JSON
自动关联人照/房照、提取组长/重点人群标签、并将所有图片以 ASCII 安全文件名复制进项目 assets。
输出: data/ledger-data.js (window.INITIAL_DATA = {...})
"""
import openpyxl, os, re, json, shutil, hashlib, datetime

# 路径自动按脚本位置推断，换电脑/换村无需改代码:
#   <ROOT>/村户慧眼台账系统/tools/生成数据_generate.py  ->  PROJ=村户慧眼台账系统, ROOT=其上级
# 也可用环境变量覆盖: VHL_ROOT(数据根目录) / VHL_XLSX(指定Excel)
HERE = os.path.dirname(os.path.abspath(__file__))
PROJ = os.path.dirname(HERE)
ROOT = os.environ.get("VHL_ROOT", os.path.dirname(PROJ))
DATA = os.path.join(ROOT, "第1期数据")
XLSX = os.environ.get("VHL_XLSX", os.path.join(DATA, "村民脱敏数据.xlsx"))

A_AERIAL   = os.path.join(PROJ, "assets", "aerial")
A_AERIAL_N = os.path.join(PROJ, "assets", "aerial_annotated")
A_PERSON   = os.path.join(PROJ, "assets", "photos", "person")
A_HOUSE    = os.path.join(PROJ, "assets", "photos", "house")
A_VILLAGE  = os.path.join(PROJ, "assets", "village")
for d in (A_AERIAL, A_AERIAL_N, A_PERSON, A_HOUSE, A_VILLAGE, os.path.join(PROJ, "data")):
    os.makedirs(d, exist_ok=True)

GROUP_ZH = {"一组": 1, "二组": 2, "三组": 3, "四组": 4, "五组": 5, "六组": 6}
ZH_NUM   = {1: "一组", 2: "二组", 3: "三组", 4: "四组", 5: "五组", 6: "六组"}
# 门牌前缀 -> 房照分组中文名
DOOR_PREFIX_TO_PHOTO = {"1": "一组", "2": "二组", "3": "三组", "5": "五组",
                        "46": "四六组", "4": "四六组", "6": "四六组"}
# 组号 -> 航拍底图 id
GROUP_TO_MAP = {1: "g1", 2: "g235", 3: "g235", 5: "g235", 4: "g46", 6: "g46"}

KEY_TAGS = [
    ("脱贫", "脱贫户"), ("低保", "低保户"), ("五保", "五保户"),
    ("残疾", "残疾人"), ("孤寡", "孤寡老人"), ("独居", "独居老人"),
    ("大病", "大病户"), ("重病", "大病户"), ("村医", "村医"),
    ("党员", "党员"), ("退役", "退役军人"), ("低收入", "低收入户"),
]

def s(v):
    return "" if v is None else str(v).strip()

# ---------- 1. 扫描照片目录 ----------
person_index = {}   # code -> [filename...]
for fn in os.listdir(os.path.join(DATA, "人房照片标注", "人照")):
    stem = os.path.splitext(fn)[0]
    person_index.setdefault(stem, []).append(fn)

house_index = {}    # (groupZh, num) -> filename
unmatched_house_files = []
for fn in os.listdir(os.path.join(DATA, "人房照片标注", "房照")):
    stem = os.path.splitext(fn)[0]
    m = re.match(r"([一二三四五六四六]+组)", stem)
    if not m:
        unmatched_house_files.append(fn); continue
    gzh = m.group(1)
    nums = re.findall(r"(\d+)", stem)
    for n in nums:
        house_index[(gzh, int(n))] = fn

# ---------- 2. 复制航拍 / 标注航拍 / 村标识 ----------
def copy_one(src, dst):
    if not os.path.exists(dst):
        shutil.copy2(src, dst)

aerial_map = {  # 原文件名 -> map id
    "一组航拍.jpg": "g1", "二三五组航拍.jpg": "g235",
    "四六组航拍.jpg": "g46", "村部航拍.jpg": "cunbu",
}
maps_meta = []
ad = os.path.join(DATA, "航拍图")
for fn, mid in aerial_map.items():
    copy_one(os.path.join(ad, fn), os.path.join(A_AERIAL, mid + ".jpg"))
annot = {"1组航拍备注门牌.jpg": "g1", "235组航拍备注门牌.jpg": "g235", "46组航拍备注门牌.jpg": "g46"}
adn = os.path.join(DATA, "航拍图-标注门牌序号")
for fn, mid in annot.items():
    p = os.path.join(adn, fn)
    if os.path.exists(p):
        copy_one(p, os.path.join(A_AERIAL_N, mid + ".jpg"))

MAP_LABELS = {"g1": "一组", "g235": "二/三/五组", "g46": "四/六组", "cunbu": "村部全景"}
for mid in ("g1", "g235", "g46", "cunbu"):
    maps_meta.append({
        "id": mid, "label": MAP_LABELS[mid],
        "image": f"assets/aerial/{mid}.jpg",
        "annotated": (f"assets/aerial_annotated/{mid}.jpg"
                      if os.path.exists(os.path.join(A_AERIAL_N, mid + ".jpg")) else "")
    })

# 村标识照片
village_imgs = []
vd = os.path.join(DATA, "村标识照片")
for i, fn in enumerate(sorted(os.listdir(vd)), 1):
    dst = f"v{i:02d}.jpg"
    copy_one(os.path.join(vd, fn), os.path.join(A_VILLAGE, dst))
    village_imgs.append(f"assets/village/{dst}")

# ---------- 3. 复制人照 (保持 code 文件名, 已是ASCII) ----------
copied_person = {}
for stem, files in person_index.items():
    out = []
    for fn in files:
        dst = fn  # ascii
        copy_one(os.path.join(DATA, "人房照片标注", "人照", fn), os.path.join(A_PERSON, dst))
        out.append(f"assets/photos/person/{dst}")
    copied_person[stem] = out

# ---------- 4. 复制房照 (中文名 -> ascii id) ----------
house_file_to_ascii = {}
hp_src = os.path.join(DATA, "人房照片标注", "房照")
for idx, fn in enumerate(sorted(os.listdir(hp_src)), 1):
    asc = f"house_{idx:03d}.jpg"
    house_file_to_ascii[fn] = asc
    copy_one(os.path.join(hp_src, fn), os.path.join(A_HOUSE, asc))

# ---------- 5. 解析 Excel ----------
wb = openpyxl.load_workbook(XLSX, data_only=True)
households = []
hid_counter = 0
group_leaders = {}
report = {"unmatched_person_codes": [], "households_without_house_photo": [],
          "members_without_photo": 0}

for ws in wb.worksheets:
    title = ws.title.strip()
    gnum = None
    for zh, n in GROUP_ZH.items():
        if zh.replace("组", "") in title or zh in title:
            gnum = n; break
    if gnum is None:
        m = re.search(r"[一二三四五六]", title)
        gnum = {"一":1,"二":2,"三":3,"四":4,"五":5,"六":6}.get(m.group(0)) if m else 0
    # 组长
    head_cell = s(ws.cell(1, 1).value)
    lm = re.search(r"组长[:：]\s*([A-Za-z0-9_]+)", head_cell)
    if lm:
        group_leaders[gnum] = lm.group(1)

    cur = None
    for r in range(3, ws.max_row + 1):
        seq   = s(ws.cell(r, 1).value)
        head  = s(ws.cell(r, 2).value)
        mem   = s(ws.cell(r, 3).value)
        rel   = s(ws.cell(r, 4).value)
        idn   = s(ws.cell(r, 5).value)
        sex   = s(ws.cell(r, 6).value)
        tel   = s(ws.cell(r, 7).value)
        door  = s(ws.cell(r, 8).value)
        hp    = s(ws.cell(r, 9).value)
        pp    = s(ws.cell(r,10).value)
        plant = s(ws.cell(r,11).value)
        breed = s(ws.cell(r,12).value)
        other = s(ws.cell(r,13).value)
        code = head or mem
        if not code and not any([rel, idn, sex, other]):
            continue
        # 电话列里偶有“去世/务工”等状态文字(非号码): 归入成员备注, 不当电话
        tel_note = ""
        if tel and not re.search(r"\d", tel):
            tel_note = tel
            tel = ""
        if seq:  # 新户
            hid_counter += 1
            cur = {
                "id": f"H{gnum}_{int(float(seq)) if re.match(r'^[0-9.]+$', seq) else seq:0>3}"
                      if re.match(r'^[0-9.]+$', seq) else f"H{gnum}_{seq}",
                "uid": f"hh-{hid_counter:04d}",
                "group": gnum, "seq": seq,
                "headCode": code, "doorplate": door,
                "phone": tel, "planting": plant, "breeding": breed,
                "houseNote": other if not code else "",
                "members": [], "housePhotos": [], "tags": [],
                "mapId": GROUP_TO_MAP.get(gnum, "cunbu"),
                "x": None, "y": None,
            }
            households.append(cur)
        if cur is None:
            continue
        if code:
            photos = copied_person.get(code, [])
            if not photos and re.match(r"^T\d+_H_", code):
                report["members_without_photo"] += 1
            note_full = "；".join([x for x in [tel_note, other] if x])
            cur["members"].append({
                "code": code, "relation": rel or ("户主" if seq else "成员"),
                "idMask": idn, "gender": sex, "note": note_full,
                "phone": tel if not seq else "", "photos": photos,
            })
        # 户级电话兜底
        if tel and not cur["phone"]:
            cur["phone"] = tel
        if plant and not cur["planting"]:
            cur["planting"] = plant
        if breed and not cur["breeding"]:
            cur["breeding"] = breed

# ---------- 6. 户级后处理: 房照关联 / 标签 / 坐标 ----------
def parse_door(door):
    if not door: return (None, None)
    m = re.match(r"\s*(\d+)\s*[-—－]\s*(\d+)", door)
    if m: return (m.group(1), int(m.group(2)))
    m2 = re.match(r"\s*(\d+)", door)
    return (m2.group(1) if m2 else None, None)

for hh in households:
    prefix, num = parse_door(hh["doorplate"])
    if prefix and num is not None:
        gzh = DOOR_PREFIX_TO_PHOTO.get(prefix)
        if gzh and (gzh, num) in house_index:
            fn = house_index[(gzh, num)]
            hh["housePhotos"] = [f"assets/photos/house/{house_file_to_ascii[fn]}"]
    if not hh["housePhotos"]:
        report["households_without_house_photo"].append(hh["id"])
    # 标签
    text = " ".join([hh.get("houseNote", "")] + [m["note"] for m in hh["members"]])
    tags = []
    for kw, tag in KEY_TAGS:
        if kw in text and tag not in tags:
            tags.append(tag)
    hh["tags"] = tags
    hh["isLeader"] = (hh["headCode"] == group_leaders.get(hh["group"]))

# 坐标: 每张底图内按户网格分布 (归一化 0~1, 可在前端拖动微调后存localStorage)
from collections import defaultdict
by_map = defaultdict(list)
for hh in households:
    by_map[hh["mapId"]].append(hh)
for mid, lst in by_map.items():
    n = len(lst)
    cols = max(1, int(round(n ** 0.5 * 1.3)))
    rows = max(1, (n + cols - 1) // cols)
    for i, hh in enumerate(lst):
        c = i % cols; rr = i // cols
        hh["x"] = round(0.07 + (0.86) * (c + 0.5) / cols, 4)
        hh["y"] = round(0.08 + (0.84) * (rr + 0.5) / rows, 4)

# ---------- 7. 汇总输出 ----------
total_members = sum(len(h["members"]) for h in households)
payload = {
    "meta": {
        "appName": "村户慧眼台账系统",
        "village": "围场县银窝沟乡某村",
        "generatedAt": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "groupLeaders": group_leaders,
        "village_images": village_imgs,
        "schemaVersion": 1,
    },
    "maps": maps_meta,
    "households": households,
}

out_js = os.path.join(PROJ, "data", "ledger-data.js")
with open(out_js, "w", encoding="utf-8") as f:
    f.write("// 自动生成: 村户脱敏台账数据 (window.INITIAL_DATA)\n")
    f.write("window.INITIAL_DATA = ")
    json.dump(payload, f, ensure_ascii=False, indent=1)
    f.write(";\n")

print("households:", len(households), "members:", total_members)
print("groups leaders:", group_leaders)
print("maps:", [m["id"] for m in maps_meta])
print("village imgs:", len(village_imgs), "house photos copied:", len(house_file_to_ascii),
      "person stems:", len(copied_person))
print("households w/o house photo:", len(report["households_without_house_photo"]))
print("members w/o photo:", report["members_without_photo"])
print("unmatched house files:", unmatched_house_files)
print("OUT:", out_js)
